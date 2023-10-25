import os
from bs4 import BeautifulSoup
import openpyxl
import re  # Import the regular expression module for pattern matching

# Directory paths


cleaned_data_folder = "Cleaned data"
excel_data_folder = "Excel data"

# Create a new Excel workbook and add a worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Define the headers for the Excel sheet
headers = [
    "Name of the Poster",
    "Time since the Post",
    "Content",
    "Hashtags",
    "Media Type",
    "External Links",
    "Number of Likes",
    "Number of Comments",
    "Number of Reposts"
]

# Write headers to the worksheet
for col_num, header in enumerate(headers, start=1):
    worksheet.cell(row=1, column=col_num, value=header)

# Function to extract data from a post and add it to the Excel worksheet
def extract_data_and_add_to_excel(post, row_num):
    name_element = post.find(class_="update-components-actor__name")
    name = name_element.get_text(strip=True).split()[0] if name_element else ""
    
    time_since_element = post.find("a", class_="app-aware-link update-components-actor__sub-description-link")
    time_since_post = time_since_element.get_text(strip=True).split("•")[0].strip() if time_since_element else ""
    
    content_element = post.find("span", class_="break-words")
    content = content_element.get_text(strip=True) if content_element else ""
    
    # Extract hashtags using regular expression from content
    hashtags = re.findall(r'#\w+', content)
    
    # Remove hashtags from content
    content = re.sub(r'#\w+', '', content).strip()
    
    external_links = [a["href"] for a in post.find_all("a", href=True) if not a["href"].startswith("https://www.linkedin.com")]
    
    # Determine the media type
    media_type = "link" if external_links else "video" if post.find("video") else "picture" if post.find("img") else "text"
    
    # Check for multiple image wrappers and update media_type if necessary
    if len(post.find_all("button", class_="update-components-image__image-link")) > 1:
        media_type = "carousel"
    
    num_likes_element = post.find(class_="social-details-social-counts__reactions-count")
    num_likes = num_likes_element.get_text(strip=True) if num_likes_element else ""
    
    num_comments_element = post.find(class_="social-details-social-counts__comments")
    num_comments = num_comments_element.get_text(strip=True).split()[0] if num_comments_element else ""
    
    num_reposts_element = post.find("button", {"aria-label": re.compile(r"\d+ reposts?")})
    num_reposts = num_reposts_element.get_text(strip=True).split()[0] if num_reposts_element else ""
    
    worksheet.cell(row=row_num, column=1, value=name)
    worksheet.cell(row=row_num, column=2, value=time_since_post)
    worksheet.cell(row=row_num, column=3, value=content)
    worksheet.cell(row=row_num, column=4, value=', '.join(hashtags))
    worksheet.cell(row=row_num, column=5, value=media_type)
    worksheet.cell(row=row_num, column=6, value=', '.join(external_links))
    worksheet.cell(row=row_num, column=7, value=num_likes)
    worksheet.cell(row=row_num, column=8, value=num_comments)
    worksheet.cell(row=row_num, column=9, value=num_reposts)


# List all cleaned files in the cleaned data folder
cleaned_files = [f for f in os.listdir(cleaned_data_folder) if f.endswith("_clean.txt")]

# Process each cleaned file
for cleaned_file in cleaned_files:
    input_cleaned_file = os.path.join(cleaned_data_folder, cleaned_file)
    # Read the cleaned data file
    with open(input_cleaned_file, "r", encoding="utf-8") as file:
        html_data = file.read()

    soup = BeautifulSoup(html_data, 'html.parser')
    posts = soup.find_all("li", class_="profile-creator-shared-feed-update__container")

    # Create a new Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Write headers to the worksheet
    for col_num, header in enumerate(headers, start=1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Initialize the row number
    row_number = 2

    # Process each post and extract data
    for post in posts:
        extract_data_and_add_to_excel(post, row_number)
        row_number += 1

    # Construct the output excel file name
    output_excel_file = os.path.join(excel_data_folder, os.path.splitext(cleaned_file)[0] + "_processed.xlsx")

    # Save the Excel file
    workbook.save(output_excel_file)

    print(f"Data extracted and saved to {output_excel_file}")

with open(input_cleaned_file, "r", encoding="utf-8") as file:
    html_data = file.read()

soup = BeautifulSoup(html_data, 'html.parser')
posts = soup.find_all("li", class_="profile-creator-shared-feed-update__container")

# Initialize the row number
row_number = 2

# Process each post and extract data
for post in posts:
    extract_data_and_add_to_excel(post, row_number)
    row_number += 1

# Save the Excel file
workbook.save(output_excel_file)

print(f"Data extracted and saved to {output_excel_file}")
